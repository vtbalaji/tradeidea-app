#!/usr/bin/env -S venv/bin/python3
"""
Extract all metrics from Raw Data sheet for validation
"""

import openpyxl
import os
from datetime import datetime

def extract_metrics(file_path, sheet_name="Quarterly - Raw Data"):
    """Extract all metrics from Raw Data sheet"""
    symbol = os.path.basename(file_path).split('-')[1].split('.')[0].upper()
    if symbol == "SBIN":
        symbol = "SBIN"
    elif symbol == "RELIANCE":
        symbol = "RELIANCE"
    else:
        symbol = symbol.upper()

    print(f"\n{'='*80}")
    print(f"📊 Extracting: {symbol} from {os.path.basename(file_path)}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        print(f"❌ Sheet '{sheet_name}' not found")
        wb.close()
        return None

    sheet = wb[sheet_name]

    # Get date headers (row 1, starting from col 2)
    dates = []
    for col_idx in range(2, min(15, sheet.max_column + 1)):
        cell = sheet.cell(1, col_idx)
        if cell.value:
            dates.append(str(cell.value))

    print(f"\n📅 Available quarters: {len(dates)}")
    print(f"  Latest: {dates[0] if dates else 'N/A'}")

    # Extract all parameters for latest quarter (column 2)
    metrics = {}
    print(f"\n📋 All Parameters (Latest Quarter):")
    print(f"  {'='*76}")

    for row_idx in range(2, sheet.max_row + 1):
        param_cell = sheet.cell(row_idx, 1)
        value_cell = sheet.cell(row_idx, 2)  # Latest quarter in column 2

        if param_cell.value:
            param_name = str(param_cell.value).strip()
            value = value_cell.value

            if value is not None and value != 0:
                metrics[param_name] = value
                print(f"  {param_name:40s} = {value:>15,.2f}")

    wb.close()

    return {
        'symbol': symbol,
        'latest_quarter': dates[0] if dates else None,
        'metrics': metrics
    }

def main():
    testdata_dir = "testdata"

    excel_files = [
        "trendlyne-TCS-v7.xlsx",
        "trendlyne-reliance-master-v7.xlsx",
        "trendlyne-sbin-v7 2.xlsx"
    ]

    results = []

    for excel_file in excel_files:
        file_path = os.path.join(testdata_dir, excel_file)
        if os.path.exists(file_path):
            try:
                data = extract_metrics(file_path)
                if data:
                    results.append(data)
            except Exception as e:
                print(f"❌ Error reading {file_path}: {e}")
                import traceback
                traceback.print_exc()

    # Summary
    print(f"\n{'='*80}")
    print(f"📈 Summary")
    print(f"{'='*80}")
    for result in results:
        print(f"\n{result['symbol']}:")
        print(f"  Quarter: {result['latest_quarter']}")
        print(f"  Metrics: {len(result['metrics'])} parameters")

if __name__ == '__main__':
    main()
